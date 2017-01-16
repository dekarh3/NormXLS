# -*- coding: utf-8 -*-
__author__ = 'Nurzhanov Edward'

# Form implementation generated from reading ui file 'normalizefield.ui'
#
# Created by: PyQt5 UI code generator 5.7

import sys
import openpyxl
from openpyxl import Workbook
from csv2xls import csv2xls

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import (QMainWindow, QFileDialog, QMessageBox, QTableWidgetItem, QComboBox)
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import QThread
import xlrd
import xlwt
import NormalizeFields as norm
from functools import partial

MANIPULATE_LABELS = ["-------------------------",
                     "ФИО_из_поля",
                     "-------------------------",
                     "ФИО_при_рождений_из_поля",
                     "-------------------------",
                     "Пол_получить_из_ФИО",
                     "Пол_подставить_свои_значения",
                     "-------------------------",
                     "Адрес_регистраций_из_поля",
                     "-------------------------",
                     "Адрес_проживания_из_поля",
                     ]

SNILS_LABEL = ["СНИЛС"]
FIO_LABELS = ["ФИО.Фамилия", "ФИО.Имя", "ФИО.Отчество"]
FIO_BIRTH_LABELS = ["ФИО_при_рождений.Фамилия", "ФИО_при_рождений.Имя", "ФИО_при_рождений.Отчество"]
GENDER_LABEL = ["Пол"]
DATE_BIRTH_LABEL = ["Дата_рождения"]
PLACE_BIRTH_LABELS = ["Место_рождения.Страна", "Место_рождения.Область", "Место_рождения.Район",
                      "Место_рождения.Город"]
PASSPORT_DATA_LABELS = ["Данные_паспорта.Серия", "Данные_паспорта.Номер", "Данные_паспорта.Дата_выдачи",
                        "Данные_паспорта.Кем_выдан", "Данные_паспорта.Код_подразделения"]
ADRESS_REG_LABELS = ["Адрес_регистраций.Индекс",
                     "Адрес_регистраций.Регион", "Адрес_регистраций.Тип_региона",
                     "Адрес_регистраций.Район", "Адрес_регистраций.Тип_района",
                     "Адрес_регистраций.Город", "Адрес_регистраций.Тип_города",
                     "Адрес_регистраций.Населенный_пункт", "Адрес_регистраций.Тип_населенного_пункта",
                     "Адрес_регистраций.Улица", "Адрес_регистраций.Тип_улицы",
                     "Адрес_регистраций.Дом",
                     "Адрес_регистраций.Корпус",
                     "Адрес_регистраций.Квартира"]

ADRESS_LIVE_LABELS = ["Адрес_проживания.Индекс",
                      "Адрес_проживания.Регион", "Адрес_проживания.Тип_региона",
                      "Адрес_проживания.Район", "Адрес_проживания.Тип_района",
                      "Адрес_проживания.Город", "Адрес_проживания.Тип_города",
                      "Адрес_проживания.Населенный_пункт", "Адрес_проживания.Тип_населенного_пункта",
                      "Адрес_проживания.Улица", "Адрес_проживания.Тип_улицы",
                      "Адрес_проживания.Дом",
                      "Адрес_проживания.Корпус",
                      "Адрес_проживания.Квартира"]

PHONES_LABELS = ["Телефон.Мобильный", "Телефон.Родственников", "Телефон.Домашний"]

#------------------------Отключил MANIPULATE_LABELS------------------------------------------------------------
# FIELDS_IN_RESULT_TABLE_COMPLETE = [SNILS_LABEL, FIO_LABELS, FIO_BIRTH_LABELS, GENDER_LABEL, DATE_BIRTH_LABEL,
#                                   PLACE_BIRTH_LABELS, PASSPORT_DATA_LABELS, ADRESS_REG_LABELS, ADRESS_LIVE_LABELS,
#                                   PHONES_LABELS, MANIPULATE_LABELS]
#------------------------Отключил MANIPULATE_LABELS------------------------------------------------------------

FIELDS_IN_RESULT_TABLE_COMPLETE = [SNILS_LABEL, FIO_LABELS, FIO_BIRTH_LABELS, GENDER_LABEL, DATE_BIRTH_LABEL,
                                   PLACE_BIRTH_LABELS, PASSPORT_DATA_LABELS, ADRESS_REG_LABELS, ADRESS_LIVE_LABELS,
                                   PHONES_LABELS]

HEAD_RESULT_EXCEL_FILE = ['СНИЛС',
                          'Фамилия', 'Имя', 'Отчество',
                          'Фамилия_при_рождении', 'Имя_при_рождении', 'Отчество_при_рождении',
                          'Пол(0_мужской,1_женский)',
                          'Дата_рождения',
                          'Страна_рождения', 'Область_рождения', 'Район_рождения', 'Город_рождения',
                          'Паспорт_серия', 'Паспорт_номер', 'Паспорт_дата', 'Паспорт_Кем выдан',
                          'Паспорт_Код подразделения',

                          'Адрес_регистрации_Индекс',
                          'Адрес_регистрации_Регион', 'Адрес_регистрации_Тип_региона',
                          'Адрес_регистрации_Район', 'Адрес_регистрации_Тип_района',
                          'Адрес_регистрации_Город', 'Адрес_регистрации_Тип_города',
                          'Адрес_регистрации_Населенный_пункт', 'Адрес_регистрации_Тип_населенного_пункта',
                          'Адрес_регистрации_Улица',
                          'Адрес_регистрации_Тип_улицы',
                          'Адрес_регистрации_Дом',
                          'Адрес_регистрации_Корпус',
                          'Адрес_регистрации_Квартира',

                          'Адрес_проживания_Индекс',
                          'Адрес_проживания_Регион', 'Адрес_проживания_Тип_региона',
                          'Адрес_проживания_Район', 'Адрес_проживания_Тип_района',
                          'Адрес_проживания_Город', 'Адрес_проживания_Тип_города',
                          'Адрес_проживания_Населенный_пункт', 'Адрес_проживания_Тип_населенного_пункта',
                          'Адрес_проживания_Улица', 'Адрес_проживания_Тип_улицы',
                          'Адрес_проживания_Дом',
                          'Адрес_проживания_Корпус',
                          'Адрес_проживания_Квартира',

                          'Мобильный_телефон', 'Телефон_родственников', 'Телефон_домашний',
                          'Агент_Ид', 'Подписант_Ид', 'Пред_Страховщик_Ид'
                          ]

class Ui_MainWindow(QMainWindow):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(800, 600)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.tableWidget = QtWidgets.QTableWidget(self.centralwidget)
        self.tableWidget.setGeometry(QtCore.QRect(160, 100, 491, 291))
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(2)

        self.tableWidget.setHorizontalHeaderLabels(('Столбец в результирующей таблице', 'Столбец в исходящей таблице'))
        for j in range(self.tableWidget.columnCount()):
            self.tableWidget.setColumnWidth(j, 220)
        self.tableWidget.setRowCount(0)

        self.Button_table_add = QtWidgets.QPushButton(self.centralwidget)
        self.Button_table_add.setGeometry(QtCore.QRect(60, 180, 50, 50))
        self.Button_table_add.setObjectName("Button_table_add")
        self.Button_table_del = QtWidgets.QPushButton(self.centralwidget)
        self.Button_table_del.setGeometry(QtCore.QRect(60, 240, 50, 50))
        self.Button_table_del.setObjectName("Button_table_del")
        self.Button_table_up = QtWidgets.QPushButton(self.centralwidget)
        self.Button_table_up.setGeometry(QtCore.QRect(670, 180, 75, 50))
        self.Button_table_up.setObjectName("Button_table_up")
        self.Button_table_down = QtWidgets.QPushButton(self.centralwidget)
        self.Button_table_down.setGeometry(QtCore.QRect(670, 240, 75, 50))
        self.Button_table_down.setObjectName("Button_table_down")
        self.Button_start = QtWidgets.QPushButton(self.centralwidget)
        self.Button_start.setGeometry(QtCore.QRect(330, 410, 161, 41))
        self.Button_start.setObjectName("Button_start")
        self.progressBar = QtWidgets.QProgressBar(self.centralwidget)
        self.progressBar.setGeometry(QtCore.QRect(100, 520, 611, 23))
        self.progressBar.setMaximum(0)
        self.progressBar.setProperty("value", 0)
        self.progressBar.setObjectName("progressBar")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 800, 21))
        self.menubar.setObjectName("menubar")
        self.menu = QtWidgets.QMenu(self.menubar)
        self.menu.setObjectName("menu")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.action = QtWidgets.QAction(MainWindow)
        self.action.setObjectName("action")
        self.menu.addAction(self.action)
        self.menubar.addAction(self.menu.menuAction())

        self.retranslateUi(MainWindow)

        self.Button_table_add.clicked.connect(self.add_table_row)
        self.Button_table_del.clicked.connect(self.del_table_row)
        self.Button_table_up.clicked.connect(partial(self.move_table_row, mode='up'))
        self.Button_table_down.clicked.connect(partial(self.move_table_row, mode='down'))
        self.Button_start.clicked.connect(self.proc)

        self.action.triggered.connect(self.FileChoise)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def proc(self):
        self.updateProgressBar(0)
#        self.FileChoise()                                              # попробовал - не работает
        self.Button_start.setEnabled(False)
        self.workerThread = WorkerThread(sheet=self.sheet,
                                         tableWidget=self.tableWidget)  #####################################
        self.workerThread.progress_value.connect(self.updateProgressBar)
        self.workerThread.start()
        self.updateProgressBar(0)
        self.Button_start.setEnabled(True)

    def updateProgressBar(self, val):
        self.progressBar.setValue(val)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Главное окно"))
        self.Button_table_add.setText(_translate("MainWindow", "+"))
        self.Button_table_del.setText(_translate("MainWindow", "-"))
        self.Button_table_up.setText(_translate("MainWindow", "up"))
        self.Button_table_down.setText(_translate("MainWindow", "down"))
        self.Button_start.setText(_translate("MainWindow", "Start"))
        self.progressBar.setFormat(_translate("MainWindow", "%v из %m"))
        self.menu.setTitle(_translate("MainWindow", "Функции"))
        self.action.setText(_translate("MainWindow", "Выбрать файл"))

    def add_table_row(self, combo1index=-1, combo2index=-1): # combo2index - индекс выбора второго QCombobox
        try:
            if self.head is None:
                pass
        except AttributeError:
            return
        self.tableWidget.insertRow(self.tableWidget.rowCount().real)
        items = []

        self.combobox_table_result = QComboBox()
        # self.combobox_table_result.setMaxVisibleItems(15)
        for row in FIELDS_IN_RESULT_TABLE_COMPLETE:
            for name in row:
                self.combobox_table_result.addItem(name)
        items.append(self.combobox_table_result)
        if combo1index != -1:
            self.combobox_table_result.setCurrentIndex(combo1index)                 # combobox_table_result - первый комбобокс
        name_combobox_table_result = "combobox_table_result_{0}".format(self.tableWidget.rowCount() - 1)
        self.combobox_table_result.setObjectName(name_combobox_table_result)

        self.combobox_table_from = QComboBox()
        # self.combobox_table_from.setMaxVisibleItems(15)
        for name in self.head:
            name = str(name)
            self.combobox_table_from.addItem(name)
        items.append(self.combobox_table_from)
        if combo2index != -1:
            self.combobox_table_from.setCurrentIndex(combo2index)                # combobox_table_from - второй комбобокс
        name_combobox_table_from = "combobox_table_from_{0}".format(self.tableWidget.rowCount() - 1)
        self.combobox_table_from.setObjectName(name_combobox_table_from)

        for n, i in enumerate(items):
            self.tableWidget.setCellWidget(self.tableWidget.rowCount() - 1, n, i)

    def del_table_row(self):
        rows = self.tableWidget.selectionModel().selectedRows()
        for i, row in enumerate(rows):
            self.tableWidget.removeRow(row.row() - i)
        self.tableWidget.clearSelection()

    def move_table_row(self, mode):
        try:
            selected_row_number = self.tableWidget.selectionModel().selectedRows()[0].row()
        except IndexError:
            return
        if mode == 'up':
            index = -1
        elif mode == 'down':
            index = 1
        else:
            index = 0
        if selected_row_number != 0:
            selected_row = [self.tableWidget.cellWidget(selected_row_number, i) for i in
                            range(self.tableWidget.columnCount())]
            new_selected_row = [self.tableWidget.cellWidget(selected_row_number + index, i) for i in
                                range(self.tableWidget.columnCount())]
            for i, item in enumerate(selected_row):
                combo = QComboBox()
                for n in range(item.count()):
                    combo.addItem(item.itemText(n))
                combo.setCurrentIndex(item.currentIndex())
                self.tableWidget.setCellWidget(selected_row_number + index, i, combo)
            for i, item in enumerate(new_selected_row):
                combo = QComboBox()
                for n in range(item.count()):
                    combo.addItem(item.itemText(n))
                combo.setCurrentIndex(item.currentIndex())
                self.tableWidget.setCellWidget(selected_row_number, i, combo)
            self.tableWidget.selectRow(selected_row_number + index)

    def FileChoise(self):
        try:
            if len(sys.argv) > 1:
                fname = sys.argv[1]
                self.fname = fname
            else:
                fname = QFileDialog.getOpenFileName(self, 'Open file', '')[0]
                self.fname = fname

#---------------------------Убрал обработку .csv файла---------------------------------------------------
#            fname = QFileDialog.getOpenFileName(self, 'Open file', '')[0]
#            self.fname = fname
#            if fname.split('.')[-1] == 'csv':
#                fname = csv2xls(fname)
# ---------------------------Убрал обработку .csv файла---------------------------------------------------

        except FileNotFoundError as f:
            for j in range(self.tableWidget.rowCount()):
                self.tableWidget.removeRow(0)
            self.progressBar.setMaximum(0)
            QMessageBox.warning(self, 'Ошибка', 'Нет файла Excel')
            return

        self.head = self.get_head_excel_file(fname)
        mass = []

        if self.sheet.max_row == None:
            for j in range(self.tableWidget.rowCount()):
                self.tableWidget.removeRow(0)
            self.progressBar.setMaximum(0)
            QMessageBox.warning(self, 'Ошибка', 'Файл Excel некорректно сохранен OpenPyxl. Откройте и пересохраните его')
            return
        else:
            self.progressBar.setMaximum(self.sheet.max_row-1)

        use_config = False
        conf_mass = []
        mass = []
        try:
            cname = fname[0:fname.rfind('xlsx')]+ 'cfg'
            conf_file = open(cname,'rt',encoding='utf-8')
            conf_file_string = conf_file.read()
            for j in range(self.tableWidget.rowCount()):
                self.tableWidget.removeRow(0)

            st = ''
            for ch in conf_file_string:                                             # заполняем conf_mass из файла cfg
                if ch != '\n':
                    st = st + ch
                else:
                    if st != '':
                        conf_mass.append(int(st))
                    st = ''
            if st!='':
                conf_mass.append(int(st))
            conf_file.close()
            use_config = True
        except FileNotFoundError as f:
            use_config = False

        for i in FIELDS_IN_RESULT_TABLE_COMPLETE:
            for name in i:
                mass.append(name)

        for i, j in enumerate(mass):
            if i < len(conf_mass) and use_config:
                if conf_mass[i] < len(mass):
                    ii = conf_mass[i]
                else:
                    ii = i                                          # по умолчанию
            else:
                ii = i                                              # по умолчанию
            self.add_table_row(i,ii)                                # добавил ii

    def get_head_excel_file(self, path, sheet_number=0):
        wb = openpyxl.load_workbook(filename=path, read_only=True)
        sheet = wb[wb.sheetnames[sheet_number]]
        self.sheet = sheet
        head = []
        for i, row in enumerate(sheet.rows):
            if i == 0:
                for cell in row:
                    head.append(cell.value)
                break
        return head


class WorkerThread(QThread):
    progress_value = QtCore.pyqtSignal(int)

    def __init__(self, tableWidget, sheet, parent=None):
        super(WorkerThread, self).__init__(parent)
        self.tableWidget = tableWidget
        self.sheet = sheet

    def run(self):
        self.start_process()

    def start_process(self):
#        fname = self.fname                                 # ????? Как передать переменную????
        fname = sys.argv[1]                                 # Без указания файла в командной строке пока не работает!!!!
        lname = fname[0:fname.rfind('xlsx')]+ 'log'
        err_from_log = {}
        use_log = False
        try:
            log_file = open(lname,'rt',encoding='utf-8')
            log_file_string = log_file.read()
            first_sq = 0
            next_str = 0
            dub_toch = 0
            last_sq = 1
            n_str_w_err = ''
            text_err = ''
            for nx in range(len(log_file_string)):
                if log_file_string[nx] == ':':
                    dub_toch = nx
                if log_file_string[nx] == '\n':
                    next_str = nx
                if log_file_string[nx] == '#' or nx == len(log_file_string) - 1:
                    first_sq = last_sq
                    last_sq = nx
                    if dub_toch > 0:
                        n_str_w_err = int(log_file_string[first_sq+1:dub_toch])
                        text_err = log_file_string[dub_toch + 3:next_str]
                        err_from_log[n_str_w_err] = text_err
            use_log = True
        except:
            use_log = False

        cname = 'new_' + fname[0:fname.rfind('xlsx')]+ 'cfg'
        conf_file = open(cname,'wt',encoding='utf-8')
        for i in range(self.tableWidget.rowCount()):
            conf_file.write(str(self.tableWidget.cellWidget(i,1).currentIndex()) + '\n')
        conf_file.close()

        wb_err = Workbook(write_only=True)
        ws_err = wb_err.create_sheet('Ошибки')
        ws_err.append(HEAD_RESULT_EXCEL_FILE)                                         # добавляем первую строку xlsx
        i10l = 0
        i10 = 0
        wb = Workbook(write_only=True)
        ws = wb.create_sheet('Лист1')
        ws.append(HEAD_RESULT_EXCEL_FILE)                                             # добавляем первую строку xlsx

        # --------------------------------------- Заменил первую строку xls файла---------------------------------------
        #        result_file_columns = [SNILS_LABEL, FIO_LABELS, FIO_BIRTH_LABELS, GENDER_LABEL, DATE_BIRTH_LABEL,
        #                            PLACE_BIRTH_LABELS, PASSPORT_DATA_LABELS, ADRESS_REG_LABELS, ADRESS_LIVE_LABELS,
        #                            PHONES_LABELS]

        #        listmerge = lambda result_file_columns: [col for label in result_file_columns for col in label] # заполняем первую строку xlsx
        #        head_result_file = listmerge(result_file_columns)


        #        ws.append(head_result_file)                                             # добавляем первую строку xlsx
        # --------------------------------------- Заменил первую строку xls файла ---------------------------------------

        for num_row, row in enumerate(self.sheet.rows):
            self.progress_value.emit(num_row + 1)  # отрисовываем ProgresBar
            if num_row == 0:
                continue
            i10 = int(num_row / 10000)
            if i10 > i10l:
                i10l = i10
                f = ui.fname.replace(ui.fname.split('/')[-1], '{0:02d}'.format(i10) + ui.fname.split('/')[-1])
                wb.save(f)
                wb = Workbook(write_only=True)
                ws = wb.create_sheet('Лист1')
                ws.append(HEAD_RESULT_EXCEL_FILE)  # добавляем первую строку xlsx
            else:

    #--------------------------------------- С этим if не добавляло первую строку ----------------------------------
    #            if num_row == 0:
    #                continue
    #--------------------------------------- С этим if не добавляло первую строку ----------------------------------

                result_row = {}

                passport = norm.Passport()
                phone = norm.Phone()

                for num_item in range(self.tableWidget.rowCount()):
                    item0 = self.tableWidget.cellWidget(num_item, 0).currentIndex()
                    item1 = self.tableWidget.cellWidget(num_item, 1).currentIndex()
                    label0 = self.tableWidget.cellWidget(num_item, 0).currentText()
                    label1 = self.tableWidget.cellWidget(num_item, 1).currentText()

                    row_item = str(row[item1].value)                         #Если преобразовывать все в стринг, то только тут
                    if row_item == 'None' or row_item == '2001-01-00' or row_item == '2001-01-00 00:00:00' \
                                          or  row_item == 'null' or  row_item == 'NULL' \
                                          or  row_item == 'заполнить' or row_item == '00.00.0000'\
                                          or row_item == '0000-00-00' or row_item == 'ERROR' \
                                          or row_item == '=#ССЫЛ!' or row_item == '#ССЫЛ!'\
                                          or row_item == '=#REF!' or row_item == '#REF!':
                        row_item = ''
                    elif row_item == '0' and label0 != 'Пол':
                        row_item = ''

                    if label0 in MANIPULATE_LABELS:

                        if label0 in ["ФИО_из_поля", "ФИО_при_рождений_из_поля"]:
                            FIO = norm.field2fio(row_item)
                            if label0 == "ФИО_из_поля":
                                lab = FIO_LABELS
                            elif label0 == "ФИО_при_рождений_из_поля":
                                lab = FIO_BIRTH_LABELS
                            for j in range(len(FIO)):
                                result_row[lab[j]] = FIO[j]
                            continue

    #------------------------------------------------------- Убрал класс Gender --------------------------------------
    #                    elif label0 == "Пол_получить_из_ФИО":
    #                        gender = norm.Gender(row_item)
    #                        result_row[GENDER_LABEL[0]] = gender.get_value()

    #                    elif label0 == "Пол_подставить_свои_значения":
    #                        gender = norm.Gender(FIO[2], gender_field_exists=True, gender=row_item) ## !!!!!!!!!!!!!!
    #                        result_row[GENDER_LABEL[0]] = gender.get_value()
    #------------------------------------------------------- Убрал класс Gender --------------------------------------

    #------------------------------------------------------- Убрал класс FullAdress ----------------------------------
    #                    elif label0 == "Адрес_регистраций_из_поля":
    #                        adress_reg = norm.FullAdress(row_item)
    #                        for z, cell in enumerate(adress_reg.get_values()):
    #                            result_row[ADRESS_REG_LABELS[z]] = cell

    #                    elif label0 == "Адрес_проживания_из_поля":
    #                        adress_zhit = norm.FullAdress(row_item)
    #                        for z, cell in enumerate(adress_zhit.get_values()):
    #                            result_row[ADRESS_LIVE_LABELS[z]] = cell
    #------------------------------------------------------- Убрал класс FullAdress ----------------------------------

                    elif label0 == '-------------------------':
                        continue
                    elif label0 in SNILS_LABEL:
                        result_row[label0] = norm.normalize_snils(row_item)
                    elif label0 in PLACE_BIRTH_LABELS:
                        result_row[label0] = row_item
                    elif label0 in PASSPORT_DATA_LABELS:
                        if PASSPORT_DATA_LABELS.index(label0) == 0:
                            passport.seriya = row_item
                        elif PASSPORT_DATA_LABELS.index(label0) == 1:
                            passport.nomer = row_item
                        elif PASSPORT_DATA_LABELS.index(label0) == 2:
                            passport.date = row_item
                        elif PASSPORT_DATA_LABELS.index(label0) == 3:
                            passport.who = norm.normalize_text(row_item)
                        elif PASSPORT_DATA_LABELS.index(label0) == 4:
                            passport.cod = row_item

                    elif label0 in PHONES_LABELS:
                        if PHONES_LABELS.index(label0) == 0:
                            phone.tel_mob = row_item
                        elif PHONES_LABELS.index(label0) == 1:
                            phone.tel_rod = row_item
                        elif PHONES_LABELS.index(label0) == 2:
                            phone.tel_dom = row_item
                    elif label0 in DATE_BIRTH_LABEL:
                        result_row[label0] = norm.normalize_date(row_item)
                    elif label0 in GENDER_LABEL:
                        result_row[label0] = norm.normalize_gender(row_item)
                    elif label0 == ADRESS_REG_LABELS[0] or label0 == ADRESS_LIVE_LABELS[0]:
                        result_row[label0] = norm.normalize_index(row_item)
                    elif label0 in ADRESS_REG_LABELS[11]:
                        result_row[label0] = norm.normalize_home(row_item)
                    elif label0 in ADRESS_LIVE_LABELS[11]:
                        result_row[label0] = norm.normalize_home(row_item)
                    elif label0 in ADRESS_REG_LABELS:
                        result_row[label0] = row_item
                    elif label0 in ADRESS_LIVE_LABELS:
                        result_row[label0] = row_item
                    else:
                        result_row[label0] = norm.normalize_text(row_item)

                for num, z in enumerate(passport.get_values()):
                    result_row[PASSPORT_DATA_LABELS[num]] = z
                for num, z in enumerate(phone.get_values()):
                    result_row[PHONES_LABELS[num]] = z

                LABELS = [SNILS_LABEL, FIO_LABELS, FIO_BIRTH_LABELS, GENDER_LABEL, DATE_BIRTH_LABEL,
                          PLACE_BIRTH_LABELS, PASSPORT_DATA_LABELS, ADRESS_REG_LABELS, ADRESS_LIVE_LABELS,
                          PHONES_LABELS]
                mass = []
                for l in LABELS:
                    for ll in l:
                        mass.append(ll)
                yum = True
                yum_phone0 = -1
                yum_phone1 = -1
                yum_phone2 = -1
                for num, cell in enumerate(mass):
                    if cell in result_row:
                        mass[num] = result_row[cell]                # заполняем mass, чтобы его добавить как строку в xlsx
                        if cell == PHONES_LABELS[0]:
                            if mass[num] == norm.ERROR_VALUE:
                                mass[num] = ''
                            yum_phone0 = num
                        elif cell == PHONES_LABELS[1]:
                            if mass[num] == norm.ERROR_VALUE:
                                mass[num] = ''
                            yum_phone1 = num
                        elif cell == PHONES_LABELS[2]:
                            if mass[num] == norm.ERROR_VALUE:
                                mass[num] = ''
                            yum_phone2 = num
                        elif mass[num] == norm.ERROR_VALUE:
                            yum = False

    #            yam = 0
    #            if len(phone.tel_mob) > 0:
    #                yam = int(phone.tel_mob)
    #            if len(phone.tel_rod) > 0:
    #                yam = yam + int(phone.tel_rod)
    #            if len(phone.tel_dom) > 0:
    #                yam = yam + int(phone.tel_dom)

                if mass[yum_phone0] == mass[yum_phone1] and mass[yum_phone0] !='':      # стираем дублирующиеся телефоны
                    mass[yum_phone1] = ''
                if mass[yum_phone1] == mass[yum_phone2] and mass[yum_phone1] !='':
                    mass[yum_phone2] = ''
                if mass[yum_phone0] == mass[yum_phone2] and mass[yum_phone0] !='':
                    mass[yum_phone2] = ''

                if mass[yum_phone0] == '' and mass[yum_phone1] == '' and mass[yum_phone2] == '':
                    yum = False                                                  # если нет ни одного телефона - ошибка


                if yum and err_from_log.get(num_row + 1) == None:
                    ws.append(mass)
    #                print(num_row, result_row['ФИО.Фамилия'], result_row['ФИО.Имя'], result_row['ФИО.Отчество'])
                else:
                    mass.append(num_row + 1)
                    mass.append(err_from_log.get(num_row + 1))
                    ws_err.append(mass)
    #                print(num_row, result_row['ФИО.Фамилия'], result_row['ФИО.Имя'], result_row['ФИО.Отчество'])
        f = ui.fname.replace(ui.fname.split('/')[-1], '{0:02d}'.format(i10+1) + ui.fname.split('/')[-1])
        wb.save(f)
        f = ui.fname.replace(ui.fname.split('/')[-1], 'err'.format(i10+1) + ui.fname.split('/')[-1])
        wb_err.save(f)
        if use_log:
            log_file.close()


class MyWindow(QMainWindow):
    def closeEvent(self, event):
        result = QMessageBox.question(self,
                                      "Выход",
                                      'Подтвердите выход',
                                      QMessageBox.Yes | QMessageBox.No)
        event.ignore()
        if result == QMessageBox.Yes:
            event.accept()


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = MyWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
